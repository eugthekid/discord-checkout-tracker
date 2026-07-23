"""
exporter.py
-----------
Takes a list of checkout rows (from the database) and turns them into a
formatted Excel workbook (.xlsx) using the openpyxl library.

WHAT openpyxl IS:
A pure-Python library for reading/writing real Excel files. No Excel install
needed. We create a Workbook, add rows to a Worksheet, and style the header.

TWO WAYS TO DELIVER THE FILE:
- export_to_buffer(): builds the workbook in memory (a BytesIO) and returns
  the bytes. The bot uploads these straight into Discord -- nothing is ever
  written to disk. This is what /export uses, so the file is downloadable from
  any device and doesn't depend on where the bot is running.
- export_to_xlsx(): saves the workbook to a file on disk instead. Kept as an
  option if you ever want a local copy.
Both share _build_workbook() so the formatting logic lives in exactly one place.
"""

import io
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import EXPORT_DIR

# The columns we show in the spreadsheet, in order, with friendly headers.
# (We skip internal columns like raw_json and created_ts.)
EXPORT_COLUMNS = [
    ("created_at_iso", "Date (UTC)"),
    ("channel_name", "Channel"),
    ("product", "Product"),
    ("profile", "Profile"),
    ("site", "Site"),
    ("module", "Module"),
    ("quantity", "Qty"),
    ("total", "Total"),
    ("total_amount", "Total ($)"),
    ("size", "Size"),
    ("mode", "Mode"),
    ("delivery", "Delivery"),
    ("payment", "Payment"),
    ("proxy_group", "Proxy Group"),
    ("checkout_id", "ID"),
    ("order_number", "Order #"),
    ("order_url", "Order URL"),
    ("is_preorder", "Is Preorder"),
]


def _make_filename(label: str) -> str:
    """Build a safe, timestamped filename so exports never collide."""
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_label = "".join(c for c in label if c.isalnum() or c in "-_") or "export"
    return f"checkouts_{safe_label}_{stamp}.xlsx"


def _build_workbook(rows: list) -> tuple[Workbook, int, float]:
    """
    Build the formatted workbook from rows. Returns (workbook, count, spend).
    This is the shared core -- neither saving to disk nor to memory lives here.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Checkouts"

    # --- Header row: bold white text on a dark fill, frozen so it stays
    #     visible while you scroll. ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F3136")  # Discord-ish dark grey
    for col_index, (_, header) in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_index, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")
    ws.freeze_panes = "A2"  # everything above row 2 (i.e. the header) stays put

    # --- Data rows ---
    total_spend = 0.0
    for row in rows:
        excel_row = []
        for key, _ in EXPORT_COLUMNS:
            excel_row.append(row[key])
        ws.append(excel_row)
        if row["total_amount"]:
            total_spend += row["total_amount"]

    # --- A summary row at the bottom ---
    ws.append([])  # blank spacer row
    summary = ["" for _ in EXPORT_COLUMNS]
    summary[0] = f"TOTAL: {len(rows)} checkouts"
    # Put the spend total under whichever column is "total_amount", so this
    # stays correct even if we add/reorder columns above.
    spend_index = [key for key, _ in EXPORT_COLUMNS].index("total_amount")
    summary[spend_index] = round(total_spend, 2)
    ws.append(summary)

    # --- Auto-size columns to fit their content (roughly) ---
    for col_index, (key, header) in enumerate(EXPORT_COLUMNS, start=1):
        longest = len(header)
        for row in rows:
            value = row[key]
            if value is not None:
                longest = max(longest, len(str(value)))
        # Cap the width so a long URL doesn't make one column enormous.
        ws.column_dimensions[get_column_letter(col_index)].width = min(longest + 2, 45)

    return wb, len(rows), round(total_spend, 2)


def export_to_buffer(rows: list, label: str = "export") -> tuple[io.BytesIO, str, int, float]:
    """
    Build the workbook entirely IN MEMORY and return
    (bytes_buffer, filename, count, total_spend).

    HOW "in memory" works: instead of a file path, openpyxl can save into a
    BytesIO -- a chunk of memory that behaves like a file. We rewind it with
    .seek(0) so whoever reads it (Discord's uploader) starts at the beginning.
    Nothing is written to any disk.
    """
    wb, count, spend = _build_workbook(rows)
    buffer = io.BytesIO()
    wb.save(buffer)      # write the .xlsx into memory, not a file
    buffer.seek(0)       # rewind to the start so it can be read back
    return buffer, _make_filename(label), count, spend


def export_to_xlsx(rows: list, label: str = "export") -> tuple[str, int, float]:
    """
    Alternative: save the workbook to a file on disk in EXPORT_DIR and return
    (file_path, count, total_spend). Not used by /export (which uploads to
    Discord), but handy if you ever want a local copy.
    """
    os.makedirs(EXPORT_DIR, exist_ok=True)
    wb, count, spend = _build_workbook(rows)
    path = os.path.join(EXPORT_DIR, _make_filename(label))
    wb.save(path)
    return path, count, spend
